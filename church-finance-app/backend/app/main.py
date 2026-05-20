from contextlib import asynccontextmanager
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import sqlite3
from typing import List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
DB_PATH = BASE_DIR / "database" / "church_finance.db"

PETTY_CASH_ALLOWED_CATEGORIES = {
    "Utilities",
    "Donation & Support",
    "Decor",
    "Telephone & Communication",
    "Repairs & Maintenance",
    "Publicity",
    "Decoration",
}


def get_connection():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def money_to_pesewas(amount: Decimal) -> int:
    amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    if amount < 0:
        raise HTTPException(status_code=400, detail="Amount cannot be negative.")

    return int(amount * 100)


def pesewas_to_money(amount_pesewas: int) -> float:
    return amount_pesewas / 100


def clean_optional_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None

    cleaned = value.strip()
    return cleaned if cleaned else None


def normalize_payment_source(source: Optional[str]) -> str:
    cleaned = (source or "bank").strip()

    if cleaned not in ["bank", "savings", "petty_cash"]:
        raise HTTPException(status_code=400, detail="Invalid payment source.")

    return cleaned


def serialize_bank(row: sqlite3.Row):
    return {
        "id": row["id"],
        "name": row["name"],
        "balance": pesewas_to_money(row["balance_pesewas"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def serialize_savings_account(row: sqlite3.Row):
    return {
        "id": row["id"],
        "name": row["name"],
        "balance": pesewas_to_money(row["balance_pesewas"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def serialize_savings_transfer(row: sqlite3.Row):
    return {
        "id": row["id"],
        "bank_id": row["bank_id"],
        "bank_name": row["bank_name"],
        "savings_account_id": row["savings_account_id"],
        "savings_account_name": row["savings_account_name"],
        "amount": pesewas_to_money(row["amount_pesewas"]),
        "note": row["note"],
        "created_at": row["created_at"],
    }


def serialize_receipt_entry(row: sqlite3.Row):
    return {
        "id": row["id"],
        "sheet_id": row["sheet_id"],
        "category": row["category"],
        "subcategory": row["subcategory"],
        "amount": pesewas_to_money(row["amount_pesewas"]),
        "bank_id": row["bank_id"],
        "bank_name": row["bank_name"],
        "note": row["note"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def serialize_receipt_sheet(row: sqlite3.Row):
    return {
        "id": row["id"],
        "title": row["title"],
        "sheet_date": row["sheet_date"],
        "status": row["status"],
        "total_amount": pesewas_to_money(row["total_pesewas"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "posted_at": row["posted_at"],
    }


def serialize_payment_entry(row: sqlite3.Row):
    return {
        "id": row["id"],
        "sheet_id": row["sheet_id"],
        "category": row["category"],
        "subcategory": row["subcategory"],
        "amount": pesewas_to_money(row["amount_pesewas"]),
        "bank_id": row["bank_id"],
        "bank_name": row["bank_name"],
        "savings_account_id": row["savings_account_id"],
        "savings_account_name": row["savings_account_name"],
        "source": row["source"],
        "note": row["note"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def serialize_payment_sheet(row: sqlite3.Row):
    return {
        "id": row["id"],
        "title": row["title"],
        "sheet_date": row["sheet_date"],
        "status": row["status"],
        "total_amount": pesewas_to_money(row["total_pesewas"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "posted_at": row["posted_at"],
    }


def serialize_petty_cash_withdrawal(row: sqlite3.Row):
    return {
        "id": row["id"],
        "bank_id": row["bank_id"],
        "bank_name": row["bank_name"],
        "amount": pesewas_to_money(row["amount_pesewas"]),
        "note": row["note"],
        "created_at": row["created_at"],
    }


def validate_bank_exists(conn: sqlite3.Connection, bank_id: int):
    bank = conn.execute(
        """
        SELECT id, name, balance_pesewas
        FROM banks
        WHERE id = ?
        """,
        (bank_id,),
    ).fetchone()

    if not bank:
        raise HTTPException(status_code=404, detail="Selected bank was not found.")

    return bank


def validate_savings_account_exists(conn: sqlite3.Connection, savings_account_id: int):
    savings_account = conn.execute(
        """
        SELECT id, name, balance_pesewas
        FROM savings_accounts
        WHERE id = ?
        """,
        (savings_account_id,),
    ).fetchone()

    if not savings_account:
        raise HTTPException(status_code=404, detail="Selected savings account was not found.")

    return savings_account


def group_amounts_by_bank(entries: List[sqlite3.Row]):
    grouped = {}

    for entry in entries:
        bank_id = entry["bank_id"]
        grouped[bank_id] = grouped.get(bank_id, 0) + entry["amount_pesewas"]

    return grouped


def group_bank_payment_amounts(entries: List[sqlite3.Row]):
    grouped = {}

    for entry in entries:
        if entry["source"] != "bank":
            continue

        bank_id = entry["bank_id"]

        if bank_id is None:
            raise HTTPException(
                status_code=400,
                detail="Bank payment entry is missing a bank.",
            )

        grouped[bank_id] = grouped.get(bank_id, 0) + entry["amount_pesewas"]

    return grouped


def group_savings_payment_amounts(entries: List[sqlite3.Row]):
    grouped = {}

    for entry in entries:
        if entry["source"] != "savings":
            continue

        savings_account_id = entry["savings_account_id"]

        if savings_account_id is None:
            raise HTTPException(
                status_code=400,
                detail="Savings payment entry is missing a savings account.",
            )

        grouped[savings_account_id] = (
            grouped.get(savings_account_id, 0) + entry["amount_pesewas"]
        )

    return grouped


def sum_petty_cash_entries(entries: List[sqlite3.Row]) -> int:
    total = 0

    for entry in entries:
        if entry["source"] == "petty_cash":
            total += entry["amount_pesewas"]

    return total


def get_petty_cash_totals(conn: sqlite3.Connection, exclude_payment_sheet_id: Optional[int] = None):
    total_withdrawn = conn.execute(
        """
        SELECT COALESCE(SUM(amount_pesewas), 0) AS total
        FROM petty_cash_withdrawals
        """
    ).fetchone()["total"]

    params = []
    exclude_clause = ""

    if exclude_payment_sheet_id is not None:
        exclude_clause = "AND payment_sheets.id != ?"
        params.append(exclude_payment_sheet_id)

    total_spent = conn.execute(
        f"""
        SELECT COALESCE(SUM(payment_entries.amount_pesewas), 0) AS total
        FROM payment_entries
        INNER JOIN payment_sheets ON payment_sheets.id = payment_entries.sheet_id
        WHERE payment_sheets.status = 'posted'
        AND payment_entries.source = 'petty_cash'
        {exclude_clause}
        """,
        params,
    ).fetchone()["total"]

    balance = total_withdrawn - total_spent

    return {
        "total_withdrawn": total_withdrawn,
        "total_spent": total_spent,
        "balance": balance,
    }


def get_petty_cash_balance_pesewas(
    conn: sqlite3.Connection,
    exclude_payment_sheet_id: Optional[int] = None,
) -> int:
    return get_petty_cash_totals(conn, exclude_payment_sheet_id)["balance"]


def ensure_payment_entries_schema(conn: sqlite3.Connection):
    existing_table = conn.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table' AND name = 'payment_entries'
        """
    ).fetchone()

    if not existing_table:
        return

    columns = conn.execute("PRAGMA table_info(payment_entries)").fetchall()
    column_map = {column["name"]: column for column in columns}
    column_names = set(column_map.keys())

    bank_id_is_not_null = bool(
        column_map.get("bank_id") and column_map["bank_id"]["notnull"] == 1
    )

    needs_rebuild = (
        bank_id_is_not_null
        or "source" not in column_names
        or "savings_account_id" not in column_names
    )

    if not needs_rebuild:
        return

    source_expression = "source" if "source" in column_names else "'bank'"
    savings_expression = (
        "savings_account_id" if "savings_account_id" in column_names else "NULL"
    )

    conn.execute("ALTER TABLE payment_entries RENAME TO payment_entries_old")

    conn.execute(
        """
        CREATE TABLE payment_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL,
            category TEXT NOT NULL,
            subcategory TEXT NOT NULL,
            amount_pesewas INTEGER NOT NULL CHECK(amount_pesewas > 0),
            bank_id INTEGER,
            savings_account_id INTEGER,
            source TEXT NOT NULL DEFAULT 'bank',
            note TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(sheet_id) REFERENCES payment_sheets(id) ON DELETE CASCADE,
            FOREIGN KEY(bank_id) REFERENCES banks(id),
            FOREIGN KEY(savings_account_id) REFERENCES savings_accounts(id)
        )
        """
    )

    conn.execute(
        f"""
        INSERT INTO payment_entries (
            id,
            sheet_id,
            category,
            subcategory,
            amount_pesewas,
            bank_id,
            savings_account_id,
            source,
            note,
            created_at,
            updated_at
        )
        SELECT
            id,
            sheet_id,
            category,
            subcategory,
            amount_pesewas,
            bank_id,
            {savings_expression},
            {source_expression},
            note,
            created_at,
            updated_at
        FROM payment_entries_old
        """
    )

    conn.execute("DROP TABLE payment_entries_old")


def get_receipt_sheet_detail(conn: sqlite3.Connection, sheet_id: int):
    sheet = conn.execute(
        """
        SELECT id, title, sheet_date, status, total_pesewas, created_at, updated_at, posted_at
        FROM receipt_sheets
        WHERE id = ?
        """,
        (sheet_id,),
    ).fetchone()

    if not sheet:
        raise HTTPException(status_code=404, detail="Receipt sheet not found.")

    entries = conn.execute(
        """
        SELECT
            receipt_entries.id,
            receipt_entries.sheet_id,
            receipt_entries.category,
            receipt_entries.subcategory,
            receipt_entries.amount_pesewas,
            receipt_entries.bank_id,
            banks.name AS bank_name,
            receipt_entries.note,
            receipt_entries.created_at,
            receipt_entries.updated_at
        FROM receipt_entries
        LEFT JOIN banks ON banks.id = receipt_entries.bank_id
        WHERE receipt_entries.sheet_id = ?
        ORDER BY receipt_entries.id ASC
        """,
        (sheet_id,),
    ).fetchall()

    data = serialize_receipt_sheet(sheet)
    data["entries"] = [serialize_receipt_entry(entry) for entry in entries]
    return data


def get_payment_sheet_detail(conn: sqlite3.Connection, sheet_id: int):
    sheet = conn.execute(
        """
        SELECT id, title, sheet_date, status, total_pesewas, created_at, updated_at, posted_at
        FROM payment_sheets
        WHERE id = ?
        """,
        (sheet_id,),
    ).fetchone()

    if not sheet:
        raise HTTPException(status_code=404, detail="Payment sheet not found.")

    entries = conn.execute(
        """
        SELECT
            payment_entries.id,
            payment_entries.sheet_id,
            payment_entries.category,
            payment_entries.subcategory,
            payment_entries.amount_pesewas,
            payment_entries.bank_id,
            banks.name AS bank_name,
            payment_entries.savings_account_id,
            savings_accounts.name AS savings_account_name,
            payment_entries.source,
            payment_entries.note,
            payment_entries.created_at,
            payment_entries.updated_at
        FROM payment_entries
        LEFT JOIN banks ON banks.id = payment_entries.bank_id
        LEFT JOIN savings_accounts ON savings_accounts.id = payment_entries.savings_account_id
        WHERE payment_entries.sheet_id = ?
        ORDER BY payment_entries.id ASC
        """,
        (sheet_id,),
    ).fetchall()

    data = serialize_payment_sheet(sheet)
    data["entries"] = [serialize_payment_entry(entry) for entry in entries]
    return data


def replace_receipt_entries(
    conn: sqlite3.Connection,
    sheet_id: int,
    entries: list,
) -> int:
    total_pesewas = 0

    conn.execute("DELETE FROM receipt_entries WHERE sheet_id = ?", (sheet_id,))

    for entry in entries:
        category = entry.category.strip()
        subcategory = entry.subcategory.strip()
        note = clean_optional_text(entry.note)
        amount_pesewas = money_to_pesewas(entry.amount)

        if not category:
            raise HTTPException(status_code=400, detail="Receipt category is required.")

        if not subcategory:
            raise HTTPException(status_code=400, detail="Receipt name is required.")

        validate_bank_exists(conn, entry.bank_id)

        conn.execute(
            """
            INSERT INTO receipt_entries (
                sheet_id,
                category,
                subcategory,
                amount_pesewas,
                bank_id,
                note
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                sheet_id,
                category,
                subcategory,
                amount_pesewas,
                entry.bank_id,
                note,
            ),
        )

        total_pesewas += amount_pesewas

    conn.execute(
        """
        UPDATE receipt_sheets
        SET total_pesewas = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (total_pesewas, sheet_id),
    )

    return total_pesewas


def replace_payment_entries(
    conn: sqlite3.Connection,
    sheet_id: int,
    entries: list,
) -> int:
    total_pesewas = 0

    conn.execute("DELETE FROM payment_entries WHERE sheet_id = ?", (sheet_id,))

    for entry in entries:
        category = entry.category.strip()
        subcategory = entry.subcategory.strip()
        note = clean_optional_text(entry.note)
        source = normalize_payment_source(entry.source)
        amount_pesewas = money_to_pesewas(entry.amount)

        if not category:
            raise HTTPException(status_code=400, detail="Payment category is required.")

        if not subcategory:
            raise HTTPException(status_code=400, detail="Payment name is required.")

        bank_id = entry.bank_id
        savings_account_id = entry.savings_account_id

        if source == "bank":
            if bank_id is None:
                raise HTTPException(
                    status_code=400,
                    detail="Bank payment entry must have a selected bank.",
                )

            validate_bank_exists(conn, bank_id)
            savings_account_id = None

        if source == "savings":
            if savings_account_id is None:
                raise HTTPException(
                    status_code=400,
                    detail="Savings payment entry must have a selected savings account.",
                )

            validate_savings_account_exists(conn, savings_account_id)
            bank_id = None

        if source == "petty_cash":
            bank_id = None
            savings_account_id = None

            if category not in PETTY_CASH_ALLOWED_CATEGORIES:
                raise HTTPException(
                    status_code=400,
                    detail=f"Petty cash cannot be used for {category}.",
                )

        conn.execute(
            """
            INSERT INTO payment_entries (
                sheet_id,
                category,
                subcategory,
                amount_pesewas,
                bank_id,
                savings_account_id,
                source,
                note
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sheet_id,
                category,
                subcategory,
                amount_pesewas,
                bank_id,
                savings_account_id,
                source,
                note,
            ),
        )

        total_pesewas += amount_pesewas

    conn.execute(
        """
        UPDATE payment_sheets
        SET total_pesewas = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (total_pesewas, sheet_id),
    )

    return total_pesewas


def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    with get_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS banks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                balance_pesewas INTEGER NOT NULL DEFAULT 0 CHECK(balance_pesewas >= 0),
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                transaction_type TEXT NOT NULL,
                category TEXT NOT NULL,
                subcategory TEXT,
                amount_pesewas INTEGER NOT NULL CHECK(amount_pesewas >= 0),
                bank_id INTEGER,
                source TEXT NOT NULL DEFAULT 'bank',
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(bank_id) REFERENCES banks(id)
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS savings_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                balance_pesewas INTEGER NOT NULL DEFAULT 0 CHECK(balance_pesewas >= 0),
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS savings_transfers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_id INTEGER NOT NULL,
                savings_account_id INTEGER NOT NULL,
                amount_pesewas INTEGER NOT NULL CHECK(amount_pesewas > 0),
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(bank_id) REFERENCES banks(id),
                FOREIGN KEY(savings_account_id) REFERENCES savings_accounts(id)
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS receipt_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                sheet_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'draft',
                total_pesewas INTEGER NOT NULL DEFAULT 0 CHECK(total_pesewas >= 0),
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                posted_at TEXT
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS receipt_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_id INTEGER NOT NULL,
                category TEXT NOT NULL,
                subcategory TEXT NOT NULL,
                amount_pesewas INTEGER NOT NULL CHECK(amount_pesewas > 0),
                bank_id INTEGER NOT NULL,
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(sheet_id) REFERENCES receipt_sheets(id) ON DELETE CASCADE,
                FOREIGN KEY(bank_id) REFERENCES banks(id)
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS payment_sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                sheet_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'draft',
                total_pesewas INTEGER NOT NULL DEFAULT 0 CHECK(total_pesewas >= 0),
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                posted_at TEXT
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS payment_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_id INTEGER NOT NULL,
                category TEXT NOT NULL,
                subcategory TEXT NOT NULL,
                amount_pesewas INTEGER NOT NULL CHECK(amount_pesewas > 0),
                bank_id INTEGER,
                savings_account_id INTEGER,
                source TEXT NOT NULL DEFAULT 'bank',
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(sheet_id) REFERENCES payment_sheets(id) ON DELETE CASCADE,
                FOREIGN KEY(bank_id) REFERENCES banks(id),
                FOREIGN KEY(savings_account_id) REFERENCES savings_accounts(id)
            )
            """
        )

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS petty_cash_withdrawals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_id INTEGER NOT NULL,
                amount_pesewas INTEGER NOT NULL CHECK(amount_pesewas > 0),
                note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(bank_id) REFERENCES banks(id)
            )
            """
        )

        for bank_name in ["NIB", "CBG", "Zenith"]:
            conn.execute(
                """
                INSERT OR IGNORE INTO banks (name, balance_pesewas)
                VALUES (?, ?)
                """,
                (bank_name, 0),
            )

        for savings_name in ["Naatoa", "Credit Union"]:
            conn.execute(
                """
                INSERT OR IGNORE INTO savings_accounts (name, balance_pesewas)
                VALUES (?, ?)
                """,
                (savings_name, 0),
            )

        ensure_payment_entries_schema(conn)

        conn.commit()


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield


app = FastAPI(
    title="Church Finance API",
    description="Local church finance and accounting backend.",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class BankCreate(BaseModel):
    name: str = Field(min_length=2)
    balance: Decimal = Field(default=Decimal("0.00"), ge=0)


class BankUpdate(BaseModel):
    name: Optional[str] = Field(default=None, min_length=2)
    balance: Optional[Decimal] = Field(default=None, ge=0)


class SavingsAccountCreate(BaseModel):
    name: str = Field(min_length=2)
    balance: Decimal = Field(default=Decimal("0.00"), ge=0)


class SavingsAccountUpdate(BaseModel):
    name: Optional[str] = Field(default=None, min_length=2)
    balance: Optional[Decimal] = Field(default=None, ge=0)


class SavingsTransferCreate(BaseModel):
    bank_id: int
    savings_account_id: int
    amount: Decimal = Field(gt=0)
    note: Optional[str] = None


class ReceiptEntryPayload(BaseModel):
    category: str = Field(min_length=2)
    subcategory: str = Field(min_length=1)
    amount: Decimal = Field(gt=0)
    bank_id: int
    note: Optional[str] = None


class ReceiptDraftPayload(BaseModel):
    sheet_id: Optional[int] = None
    title: str = Field(min_length=2)
    entries: List[ReceiptEntryPayload] = Field(default_factory=list)


class ReceiptPostedUpdatePayload(BaseModel):
    title: str = Field(min_length=2)
    entries: List[ReceiptEntryPayload] = Field(default_factory=list)


class PaymentEntryPayload(BaseModel):
    category: str = Field(min_length=2)
    subcategory: str = Field(min_length=1)
    amount: Decimal = Field(gt=0)
    bank_id: Optional[int] = None
    savings_account_id: Optional[int] = None
    source: str = Field(default="bank")
    note: Optional[str] = None


class PaymentDraftPayload(BaseModel):
    sheet_id: Optional[int] = None
    title: str = Field(min_length=2)
    entries: List[PaymentEntryPayload] = Field(default_factory=list)


class PaymentPostedUpdatePayload(BaseModel):
    title: str = Field(min_length=2)
    entries: List[PaymentEntryPayload] = Field(default_factory=list)


class PettyCashWithdrawalCreate(BaseModel):
    bank_id: int
    amount: Decimal = Field(gt=0)
    note: Optional[str] = None


@app.get("/api/health")
def health_check():
    return {
        "ok": True,
        "service": "Church Finance API",
        "database": str(DB_PATH),
    }


@app.get("/api/banks")
def get_banks():
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, name, balance_pesewas, created_at, updated_at
            FROM banks
            ORDER BY name ASC
            """
        ).fetchall()

    return [serialize_bank(row) for row in rows]


@app.post("/api/banks")
def create_bank(payload: BankCreate):
    bank_name = payload.name.strip()

    if not bank_name:
        raise HTTPException(status_code=400, detail="Bank name is required.")

    balance_pesewas = money_to_pesewas(payload.balance)

    try:
        with get_connection() as conn:
            cursor = conn.execute(
                """
                INSERT INTO banks (name, balance_pesewas)
                VALUES (?, ?)
                """,
                (bank_name, balance_pesewas),
            )
            conn.commit()

            row = conn.execute(
                """
                SELECT id, name, balance_pesewas, created_at, updated_at
                FROM banks
                WHERE id = ?
                """,
                (cursor.lastrowid,),
            ).fetchone()

            return serialize_bank(row)

    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="This bank already exists.")


@app.patch("/api/banks/{bank_id}")
def update_bank(bank_id: int, payload: BankUpdate):
    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id, name, balance_pesewas, created_at, updated_at
            FROM banks
            WHERE id = ?
            """,
            (bank_id,),
        ).fetchone()

        if not existing:
            raise HTTPException(status_code=404, detail="Bank not found.")

        new_name = existing["name"]
        new_balance = existing["balance_pesewas"]

        if payload.name is not None:
            new_name = payload.name.strip()
            if not new_name:
                raise HTTPException(status_code=400, detail="Bank name is required.")

        if payload.balance is not None:
            new_balance = money_to_pesewas(payload.balance)

        try:
            conn.execute(
                """
                UPDATE banks
                SET name = ?, balance_pesewas = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (new_name, new_balance, bank_id),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="This bank already exists.")

        updated = conn.execute(
            """
            SELECT id, name, balance_pesewas, created_at, updated_at
            FROM banks
            WHERE id = ?
            """,
            (bank_id,),
        ).fetchone()

    return serialize_bank(updated)


@app.delete("/api/banks/{bank_id}")
def delete_bank(bank_id: int):
    with get_connection() as conn:
        existing = conn.execute("SELECT id FROM banks WHERE id = ?", (bank_id,)).fetchone()

        if not existing:
            raise HTTPException(status_code=404, detail="Bank not found.")

        transaction_count = conn.execute(
            "SELECT COUNT(*) AS total FROM transactions WHERE bank_id = ?",
            (bank_id,),
        ).fetchone()["total"]

        receipt_entry_count = conn.execute(
            "SELECT COUNT(*) AS total FROM receipt_entries WHERE bank_id = ?",
            (bank_id,),
        ).fetchone()["total"]

        payment_entry_count = conn.execute(
            "SELECT COUNT(*) AS total FROM payment_entries WHERE bank_id = ?",
            (bank_id,),
        ).fetchone()["total"]

        savings_transfer_count = conn.execute(
            "SELECT COUNT(*) AS total FROM savings_transfers WHERE bank_id = ?",
            (bank_id,),
        ).fetchone()["total"]

        petty_cash_count = conn.execute(
            "SELECT COUNT(*) AS total FROM petty_cash_withdrawals WHERE bank_id = ?",
            (bank_id,),
        ).fetchone()["total"]

        if (
            transaction_count > 0
            or receipt_entry_count > 0
            or payment_entry_count > 0
            or savings_transfer_count > 0
            or petty_cash_count > 0
        ):
            raise HTTPException(
                status_code=400,
                detail="This bank has records and cannot be deleted.",
            )

        conn.execute("DELETE FROM banks WHERE id = ?", (bank_id,))
        conn.commit()

    return {"ok": True, "message": "Bank deleted successfully."}


@app.get("/api/savings/accounts")
def get_savings_accounts():
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, name, balance_pesewas, created_at, updated_at
            FROM savings_accounts
            ORDER BY name ASC
            """
        ).fetchall()

    return [serialize_savings_account(row) for row in rows]


@app.post("/api/savings/accounts")
def create_savings_account(payload: SavingsAccountCreate):
    name = payload.name.strip()

    if not name:
        raise HTTPException(status_code=400, detail="Savings account name is required.")

    balance_pesewas = money_to_pesewas(payload.balance)

    try:
        with get_connection() as conn:
            cursor = conn.execute(
                """
                INSERT INTO savings_accounts (name, balance_pesewas)
                VALUES (?, ?)
                """,
                (name, balance_pesewas),
            )
            conn.commit()

            row = conn.execute(
                """
                SELECT id, name, balance_pesewas, created_at, updated_at
                FROM savings_accounts
                WHERE id = ?
                """,
                (cursor.lastrowid,),
            ).fetchone()

            return serialize_savings_account(row)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="This savings account already exists.")


@app.patch("/api/savings/accounts/{savings_account_id}")
def update_savings_account(savings_account_id: int, payload: SavingsAccountUpdate):
    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id, name, balance_pesewas, created_at, updated_at
            FROM savings_accounts
            WHERE id = ?
            """,
            (savings_account_id,),
        ).fetchone()

        if not existing:
            raise HTTPException(status_code=404, detail="Savings account not found.")

        new_name = existing["name"]
        new_balance = existing["balance_pesewas"]

        if payload.name is not None:
            new_name = payload.name.strip()
            if not new_name:
                raise HTTPException(status_code=400, detail="Savings account name is required.")

        if payload.balance is not None:
            new_balance = money_to_pesewas(payload.balance)

        try:
            conn.execute(
                """
                UPDATE savings_accounts
                SET name = ?,
                    balance_pesewas = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (new_name, new_balance, savings_account_id),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="This savings account already exists.")

        updated = conn.execute(
            """
            SELECT id, name, balance_pesewas, created_at, updated_at
            FROM savings_accounts
            WHERE id = ?
            """,
            (savings_account_id,),
        ).fetchone()

    return serialize_savings_account(updated)


@app.delete("/api/savings/accounts/{savings_account_id}")
def delete_savings_account(savings_account_id: int):
    with get_connection() as conn:
        existing = conn.execute(
            """
            SELECT id
            FROM savings_accounts
            WHERE id = ?
            """,
            (savings_account_id,),
        ).fetchone()

        if not existing:
            raise HTTPException(status_code=404, detail="Savings account not found.")

        transfer_count = conn.execute(
            """
            SELECT COUNT(*) AS total
            FROM savings_transfers
            WHERE savings_account_id = ?
            """,
            (savings_account_id,),
        ).fetchone()["total"]

        payment_count = conn.execute(
            """
            SELECT COUNT(*) AS total
            FROM payment_entries
            WHERE savings_account_id = ?
            """,
            (savings_account_id,),
        ).fetchone()["total"]

        if transfer_count > 0 or payment_count > 0:
            raise HTTPException(
                status_code=400,
                detail="This savings account has records and cannot be deleted.",
            )

        conn.execute(
            """
            DELETE FROM savings_accounts
            WHERE id = ?
            """,
            (savings_account_id,),
        )
        conn.commit()

    return {"ok": True, "message": "Savings account deleted successfully."}


@app.get("/api/savings/transfers")
def get_savings_transfers():
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                savings_transfers.id,
                savings_transfers.bank_id,
                banks.name AS bank_name,
                savings_transfers.savings_account_id,
                savings_accounts.name AS savings_account_name,
                savings_transfers.amount_pesewas,
                savings_transfers.note,
                savings_transfers.created_at
            FROM savings_transfers
            INNER JOIN banks ON banks.id = savings_transfers.bank_id
            INNER JOIN savings_accounts ON savings_accounts.id = savings_transfers.savings_account_id
            ORDER BY savings_transfers.created_at DESC, savings_transfers.id DESC
            """
        ).fetchall()

    return [serialize_savings_transfer(row) for row in rows]


@app.post("/api/savings/transfers")
def create_savings_transfer(payload: SavingsTransferCreate):
    amount_pesewas = money_to_pesewas(payload.amount)
    note = clean_optional_text(payload.note)

    with get_connection() as conn:
        bank = validate_bank_exists(conn, payload.bank_id)
        savings_account = validate_savings_account_exists(conn, payload.savings_account_id)

        if bank["balance_pesewas"] - amount_pesewas < 0:
            raise HTTPException(
                status_code=400,
                detail=f"{bank['name']} does not have enough balance for this savings transfer.",
            )

        conn.execute(
            """
            UPDATE banks
            SET balance_pesewas = balance_pesewas - ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (amount_pesewas, payload.bank_id),
        )

        conn.execute(
            """
            UPDATE savings_accounts
            SET balance_pesewas = balance_pesewas + ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (amount_pesewas, payload.savings_account_id),
        )

        cursor = conn.execute(
            """
            INSERT INTO savings_transfers (
                bank_id,
                savings_account_id,
                amount_pesewas,
                note
            )
            VALUES (?, ?, ?, ?)
            """,
            (
                payload.bank_id,
                payload.savings_account_id,
                amount_pesewas,
                note,
            ),
        )

        conn.commit()

        row = conn.execute(
            """
            SELECT
                savings_transfers.id,
                savings_transfers.bank_id,
                banks.name AS bank_name,
                savings_transfers.savings_account_id,
                savings_accounts.name AS savings_account_name,
                savings_transfers.amount_pesewas,
                savings_transfers.note,
                savings_transfers.created_at
            FROM savings_transfers
            INNER JOIN banks ON banks.id = savings_transfers.bank_id
            INNER JOIN savings_accounts ON savings_accounts.id = savings_transfers.savings_account_id
            WHERE savings_transfers.id = ?
            """,
            (cursor.lastrowid,),
        ).fetchone()

    return serialize_savings_transfer(row)


@app.delete("/api/savings/transfers/{transfer_id}")
def delete_savings_transfer(transfer_id: int):
    with get_connection() as conn:
        transfer = conn.execute(
            """
            SELECT id, bank_id, savings_account_id, amount_pesewas
            FROM savings_transfers
            WHERE id = ?
            """,
            (transfer_id,),
        ).fetchone()

        if not transfer:
            raise HTTPException(status_code=404, detail="Savings transfer not found.")

        savings_account = validate_savings_account_exists(
            conn,
            transfer["savings_account_id"],
        )

        if savings_account["balance_pesewas"] - transfer["amount_pesewas"] < 0:
            raise HTTPException(
                status_code=400,
                detail="Cannot delete this transfer because part of the savings has already been used.",
            )

        conn.execute(
            """
            UPDATE banks
            SET balance_pesewas = balance_pesewas + ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (transfer["amount_pesewas"], transfer["bank_id"]),
        )

        conn.execute(
            """
            UPDATE savings_accounts
            SET balance_pesewas = balance_pesewas - ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (transfer["amount_pesewas"], transfer["savings_account_id"]),
        )

        conn.execute(
            """
            DELETE FROM savings_transfers
            WHERE id = ?
            """,
            (transfer_id,),
        )

        conn.commit()

    return {"ok": True, "message": "Savings transfer deleted successfully."}
@app.get("/api/receipt-sheets")
def get_receipt_sheets():
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, title, sheet_date, status, total_pesewas, created_at, updated_at, posted_at
            FROM receipt_sheets
            ORDER BY sheet_date DESC, id DESC
            """
        ).fetchall()

    return [serialize_receipt_sheet(row) for row in rows]


@app.get("/api/receipt-sheets/draft")
def get_latest_receipt_draft():
    with get_connection() as conn:
        draft = conn.execute(
            """
            SELECT id
            FROM receipt_sheets
            WHERE status = 'draft'
            ORDER BY id DESC
            LIMIT 1
            """
        ).fetchone()

        if not draft:
            return None

        return get_receipt_sheet_detail(conn, draft["id"])


@app.post("/api/receipt-sheets/draft")
def save_receipt_draft(payload: ReceiptDraftPayload):
    title = payload.title.strip()

    if not title:
        raise HTTPException(status_code=400, detail="Receipt sheet title is required.")

    with get_connection() as conn:
        if payload.sheet_id:
            sheet = conn.execute(
                "SELECT id, status FROM receipt_sheets WHERE id = ?",
                (payload.sheet_id,),
            ).fetchone()

            if not sheet:
                raise HTTPException(status_code=404, detail="Receipt sheet not found.")

            if sheet["status"] != "draft":
                raise HTTPException(
                    status_code=400,
                    detail="Only draft receipt sheets can be auto-saved.",
                )

            sheet_id = payload.sheet_id

            conn.execute(
                """
                UPDATE receipt_sheets
                SET title = ?,
                    sheet_date = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (title, date.today().isoformat(), sheet_id),
            )
        else:
            cursor = conn.execute(
                """
                INSERT INTO receipt_sheets (title, sheet_date, status, total_pesewas)
                VALUES (?, ?, 'draft', 0)
                """,
                (title, date.today().isoformat()),
            )
            sheet_id = cursor.lastrowid

        replace_receipt_entries(conn, sheet_id, payload.entries)

        conn.commit()

        return get_receipt_sheet_detail(conn, sheet_id)


@app.get("/api/receipt-sheets/{sheet_id}")
def get_receipt_sheet(sheet_id: int):
    with get_connection() as conn:
        return get_receipt_sheet_detail(conn, sheet_id)


@app.post("/api/receipt-sheets/{sheet_id}/post")
def post_receipt_sheet(sheet_id: int):
    with get_connection() as conn:
        sheet = conn.execute(
            "SELECT id, status FROM receipt_sheets WHERE id = ?",
            (sheet_id,),
        ).fetchone()

        if not sheet:
            raise HTTPException(status_code=404, detail="Receipt sheet not found.")

        if sheet["status"] == "posted":
            raise HTTPException(
                status_code=400,
                detail="This receipt sheet has already been posted.",
            )

        entries = conn.execute(
            """
            SELECT id, bank_id, amount_pesewas
            FROM receipt_entries
            WHERE sheet_id = ?
            """,
            (sheet_id,),
        ).fetchall()

        if len(entries) == 0:
            raise HTTPException(
                status_code=400,
                detail="Cannot post an empty receipt sheet.",
            )

        for entry in entries:
            validate_bank_exists(conn, entry["bank_id"])
            conn.execute(
                """
                UPDATE banks
                SET balance_pesewas = balance_pesewas + ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (entry["amount_pesewas"], entry["bank_id"]),
            )

        conn.execute(
            """
            UPDATE receipt_sheets
            SET status = 'posted',
                posted_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (sheet_id,),
        )

        conn.commit()

        return get_receipt_sheet_detail(conn, sheet_id)


@app.patch("/api/receipt-sheets/{sheet_id}")
def update_posted_receipt_sheet(sheet_id: int, payload: ReceiptPostedUpdatePayload):
    title = payload.title.strip()

    if not title:
        raise HTTPException(status_code=400, detail="Receipt sheet title is required.")

    with get_connection() as conn:
        sheet = conn.execute(
            "SELECT id, status FROM receipt_sheets WHERE id = ?",
            (sheet_id,),
        ).fetchone()

        if not sheet:
            raise HTTPException(status_code=404, detail="Receipt sheet not found.")

        if sheet["status"] != "posted":
            raise HTTPException(
                status_code=400,
                detail="Only posted sheets should be updated here.",
            )

        old_entries = conn.execute(
            """
            SELECT id, bank_id, amount_pesewas
            FROM receipt_entries
            WHERE sheet_id = ?
            """,
            (sheet_id,),
        ).fetchall()

        old_bank_groups = group_amounts_by_bank(old_entries)

        for bank_id, amount_pesewas in old_bank_groups.items():
            bank = validate_bank_exists(conn, bank_id)

            if bank["balance_pesewas"] - amount_pesewas < 0:
                raise HTTPException(
                    status_code=400,
                    detail="This edit cannot be completed because reversing the old receipt would make a bank balance negative.",
                )

        for bank_id, amount_pesewas in old_bank_groups.items():
            conn.execute(
                """
                UPDATE banks
                SET balance_pesewas = balance_pesewas - ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (amount_pesewas, bank_id),
            )

        conn.execute(
            """
            UPDATE receipt_sheets
            SET title = ?,
                sheet_date = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (title, date.today().isoformat(), sheet_id),
        )

        replace_receipt_entries(conn, sheet_id, payload.entries)

        new_entries = conn.execute(
            """
            SELECT id, bank_id, amount_pesewas
            FROM receipt_entries
            WHERE sheet_id = ?
            """,
            (sheet_id,),
        ).fetchall()

        for entry in new_entries:
            conn.execute(
                """
                UPDATE banks
                SET balance_pesewas = balance_pesewas + ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (entry["amount_pesewas"], entry["bank_id"]),
            )

        conn.commit()

        return get_receipt_sheet_detail(conn, sheet_id)


@app.delete("/api/receipt-sheets/{sheet_id}")
def delete_receipt_sheet(sheet_id: int):
    with get_connection() as conn:
        sheet = conn.execute(
            "SELECT id, status FROM receipt_sheets WHERE id = ?",
            (sheet_id,),
        ).fetchone()

        if not sheet:
            raise HTTPException(status_code=404, detail="Receipt sheet not found.")

        entries = conn.execute(
            """
            SELECT id, bank_id, amount_pesewas
            FROM receipt_entries
            WHERE sheet_id = ?
            """,
            (sheet_id,),
        ).fetchall()

        if sheet["status"] == "posted":
            bank_groups = group_amounts_by_bank(entries)

            for bank_id, amount_pesewas in bank_groups.items():
                bank = validate_bank_exists(conn, bank_id)

                if bank["balance_pesewas"] - amount_pesewas < 0:
                    raise HTTPException(
                        status_code=400,
                        detail="Cannot delete this sheet because reversing it would make a bank balance negative.",
                    )

            for bank_id, amount_pesewas in bank_groups.items():
                conn.execute(
                    """
                    UPDATE banks
                    SET balance_pesewas = balance_pesewas - ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (amount_pesewas, bank_id),
                )

        conn.execute("DELETE FROM receipt_entries WHERE sheet_id = ?", (sheet_id,))
        conn.execute("DELETE FROM receipt_sheets WHERE id = ?", (sheet_id,))
        conn.commit()

    return {"ok": True, "message": "Receipt sheet deleted successfully."}


@app.get("/api/petty-cash/summary")
def get_petty_cash_summary():
    with get_connection() as conn:
        totals = get_petty_cash_totals(conn)

    return {
        "balance": pesewas_to_money(totals["balance"]),
        "total_withdrawn": pesewas_to_money(totals["total_withdrawn"]),
        "total_spent": pesewas_to_money(totals["total_spent"]),
    }


@app.get("/api/petty-cash/withdrawals")
def get_petty_cash_withdrawals():
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                petty_cash_withdrawals.id,
                petty_cash_withdrawals.bank_id,
                banks.name AS bank_name,
                petty_cash_withdrawals.amount_pesewas,
                petty_cash_withdrawals.note,
                petty_cash_withdrawals.created_at
            FROM petty_cash_withdrawals
            INNER JOIN banks ON banks.id = petty_cash_withdrawals.bank_id
            ORDER BY petty_cash_withdrawals.created_at DESC, petty_cash_withdrawals.id DESC
            """
        ).fetchall()

    return [serialize_petty_cash_withdrawal(row) for row in rows]


@app.post("/api/petty-cash/withdrawals")
def create_petty_cash_withdrawal(payload: PettyCashWithdrawalCreate):
    amount_pesewas = money_to_pesewas(payload.amount)
    note = clean_optional_text(payload.note)

    with get_connection() as conn:
        bank = validate_bank_exists(conn, payload.bank_id)

        if bank["balance_pesewas"] - amount_pesewas < 0:
            raise HTTPException(
                status_code=400,
                detail=f"{bank['name']} does not have enough balance for this petty cash withdrawal.",
            )

        conn.execute(
            """
            UPDATE banks
            SET balance_pesewas = balance_pesewas - ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (amount_pesewas, payload.bank_id),
        )

        cursor = conn.execute(
            """
            INSERT INTO petty_cash_withdrawals (bank_id, amount_pesewas, note)
            VALUES (?, ?, ?)
            """,
            (payload.bank_id, amount_pesewas, note),
        )

        conn.commit()

        row = conn.execute(
            """
            SELECT
                petty_cash_withdrawals.id,
                petty_cash_withdrawals.bank_id,
                banks.name AS bank_name,
                petty_cash_withdrawals.amount_pesewas,
                petty_cash_withdrawals.note,
                petty_cash_withdrawals.created_at
            FROM petty_cash_withdrawals
            INNER JOIN banks ON banks.id = petty_cash_withdrawals.bank_id
            WHERE petty_cash_withdrawals.id = ?
            """,
            (cursor.lastrowid,),
        ).fetchone()

    return serialize_petty_cash_withdrawal(row)


@app.delete("/api/petty-cash/withdrawals/{withdrawal_id}")
def delete_petty_cash_withdrawal(withdrawal_id: int):
    with get_connection() as conn:
        withdrawal = conn.execute(
            """
            SELECT id, bank_id, amount_pesewas
            FROM petty_cash_withdrawals
            WHERE id = ?
            """,
            (withdrawal_id,),
        ).fetchone()

        if not withdrawal:
            raise HTTPException(status_code=404, detail="Petty cash withdrawal not found.")

        current_balance = get_petty_cash_balance_pesewas(conn)

        if current_balance - withdrawal["amount_pesewas"] < 0:
            raise HTTPException(
                status_code=400,
                detail="Cannot delete this withdrawal because some of the petty cash has already been spent.",
            )

        conn.execute(
            """
            UPDATE banks
            SET balance_pesewas = balance_pesewas + ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (withdrawal["amount_pesewas"], withdrawal["bank_id"]),
        )

        conn.execute(
            "DELETE FROM petty_cash_withdrawals WHERE id = ?",
            (withdrawal_id,),
        )

        conn.commit()

    return {"ok": True, "message": "Petty cash withdrawal deleted successfully."}
@app.get("/api/payment-sheets")
def get_payment_sheets():
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, title, sheet_date, status, total_pesewas, created_at, updated_at, posted_at
            FROM payment_sheets
            ORDER BY sheet_date DESC, id DESC
            """
        ).fetchall()

    return [serialize_payment_sheet(row) for row in rows]


@app.get("/api/payment-sheets/draft")
def get_latest_payment_draft():
    with get_connection() as conn:
        draft = conn.execute(
            """
            SELECT id
            FROM payment_sheets
            WHERE status = 'draft'
            ORDER BY id DESC
            LIMIT 1
            """
        ).fetchone()

        if not draft:
            return None

        return get_payment_sheet_detail(conn, draft["id"])


@app.post("/api/payment-sheets/draft")
def save_payment_draft(payload: PaymentDraftPayload):
    title = payload.title.strip()

    if not title:
        raise HTTPException(status_code=400, detail="Payment sheet title is required.")

    with get_connection() as conn:
        if payload.sheet_id:
            sheet = conn.execute(
                "SELECT id, status FROM payment_sheets WHERE id = ?",
                (payload.sheet_id,),
            ).fetchone()

            if not sheet:
                raise HTTPException(status_code=404, detail="Payment sheet not found.")

            if sheet["status"] != "draft":
                raise HTTPException(
                    status_code=400,
                    detail="Only draft payment sheets can be auto-saved.",
                )

            sheet_id = payload.sheet_id

            conn.execute(
                """
                UPDATE payment_sheets
                SET title = ?,
                    sheet_date = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (title, date.today().isoformat(), sheet_id),
            )
        else:
            cursor = conn.execute(
                """
                INSERT INTO payment_sheets (title, sheet_date, status, total_pesewas)
                VALUES (?, ?, 'draft', 0)
                """,
                (title, date.today().isoformat()),
            )
            sheet_id = cursor.lastrowid

        replace_payment_entries(conn, sheet_id, payload.entries)

        conn.commit()

        return get_payment_sheet_detail(conn, sheet_id)


@app.get("/api/payment-sheets/{sheet_id}")
def get_payment_sheet(sheet_id: int):
    with get_connection() as conn:
        return get_payment_sheet_detail(conn, sheet_id)


@app.post("/api/payment-sheets/{sheet_id}/post")
def post_payment_sheet(sheet_id: int):
    with get_connection() as conn:
        sheet = conn.execute(
            "SELECT id, status FROM payment_sheets WHERE id = ?",
            (sheet_id,),
        ).fetchone()

        if not sheet:
            raise HTTPException(status_code=404, detail="Payment sheet not found.")

        if sheet["status"] == "posted":
            raise HTTPException(
                status_code=400,
                detail="This payment sheet has already been posted.",
            )

        entries = conn.execute(
            """
            SELECT id, bank_id, savings_account_id, amount_pesewas, source
            FROM payment_entries
            WHERE sheet_id = ?
            """,
            (sheet_id,),
        ).fetchall()

        if len(entries) == 0:
            raise HTTPException(
                status_code=400,
                detail="Cannot post an empty payment sheet.",
            )

        bank_groups = group_bank_payment_amounts(entries)
        savings_groups = group_savings_payment_amounts(entries)
        petty_cash_total = sum_petty_cash_entries(entries)

        for bank_id, amount_pesewas in bank_groups.items():
            bank = validate_bank_exists(conn, bank_id)

            if bank["balance_pesewas"] - amount_pesewas < 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot post payment sheet. {bank['name']} does not have enough balance.",
                )

        for savings_account_id, amount_pesewas in savings_groups.items():
            savings_account = validate_savings_account_exists(conn, savings_account_id)

            if savings_account["balance_pesewas"] - amount_pesewas < 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot post payment sheet. {savings_account['name']} does not have enough balance.",
                )

        petty_cash_balance = get_petty_cash_balance_pesewas(conn)

        if petty_cash_balance - petty_cash_total < 0:
            raise HTTPException(
                status_code=400,
                detail="Cannot post payment sheet. Petty cash balance is not enough.",
            )

        for bank_id, amount_pesewas in bank_groups.items():
            conn.execute(
                """
                UPDATE banks
                SET balance_pesewas = balance_pesewas - ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (amount_pesewas, bank_id),
            )

        for savings_account_id, amount_pesewas in savings_groups.items():
            conn.execute(
                """
                UPDATE savings_accounts
                SET balance_pesewas = balance_pesewas - ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (amount_pesewas, savings_account_id),
            )

        conn.execute(
            """
            UPDATE payment_sheets
            SET status = 'posted',
                posted_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (sheet_id,),
        )

        conn.commit()

        return get_payment_sheet_detail(conn, sheet_id)


@app.patch("/api/payment-sheets/{sheet_id}")
def update_posted_payment_sheet(sheet_id: int, payload: PaymentPostedUpdatePayload):
    title = payload.title.strip()

    if not title:
        raise HTTPException(status_code=400, detail="Payment sheet title is required.")

    with get_connection() as conn:
        sheet = conn.execute(
            "SELECT id, status FROM payment_sheets WHERE id = ?",
            (sheet_id,),
        ).fetchone()

        if not sheet:
            raise HTTPException(status_code=404, detail="Payment sheet not found.")

        if sheet["status"] != "posted":
            raise HTTPException(
                status_code=400,
                detail="Only posted sheets should be updated here.",
            )

        old_entries = conn.execute(
            """
            SELECT id, bank_id, savings_account_id, amount_pesewas, source
            FROM payment_entries
            WHERE sheet_id = ?
            """,
            (sheet_id,),
        ).fetchall()

        old_bank_groups = group_bank_payment_amounts(old_entries)
        old_savings_groups = group_savings_payment_amounts(old_entries)

        for bank_id, amount_pesewas in old_bank_groups.items():
            validate_bank_exists(conn, bank_id)
            conn.execute(
                """
                UPDATE banks
                SET balance_pesewas = balance_pesewas + ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (amount_pesewas, bank_id),
            )

        for savings_account_id, amount_pesewas in old_savings_groups.items():
            validate_savings_account_exists(conn, savings_account_id)
            conn.execute(
                """
                UPDATE savings_accounts
                SET balance_pesewas = balance_pesewas + ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (amount_pesewas, savings_account_id),
            )

        conn.execute(
            """
            UPDATE payment_sheets
            SET title = ?,
                sheet_date = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (title, date.today().isoformat(), sheet_id),
        )

        replace_payment_entries(conn, sheet_id, payload.entries)

        new_entries = conn.execute(
            """
            SELECT id, bank_id, savings_account_id, amount_pesewas, source
            FROM payment_entries
            WHERE sheet_id = ?
            """,
            (sheet_id,),
        ).fetchall()

        new_bank_groups = group_bank_payment_amounts(new_entries)
        new_savings_groups = group_savings_payment_amounts(new_entries)
        new_petty_cash_total = sum_petty_cash_entries(new_entries)

        for bank_id, amount_pesewas in new_bank_groups.items():
            bank = validate_bank_exists(conn, bank_id)

            if bank["balance_pesewas"] - amount_pesewas < 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"This edit cannot be saved. {bank['name']} does not have enough balance.",
                )

        for savings_account_id, amount_pesewas in new_savings_groups.items():
            savings_account = validate_savings_account_exists(conn, savings_account_id)

            if savings_account["balance_pesewas"] - amount_pesewas < 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"This edit cannot be saved. {savings_account['name']} does not have enough balance.",
                )

        available_petty_cash = get_petty_cash_balance_pesewas(
            conn,
            exclude_payment_sheet_id=sheet_id,
        )

        if available_petty_cash - new_petty_cash_total < 0:
            raise HTTPException(
                status_code=400,
                detail="This edit cannot be saved. Petty cash balance is not enough.",
            )

        for bank_id, amount_pesewas in new_bank_groups.items():
            conn.execute(
                """
                UPDATE banks
                SET balance_pesewas = balance_pesewas - ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (amount_pesewas, bank_id),
            )

        for savings_account_id, amount_pesewas in new_savings_groups.items():
            conn.execute(
                """
                UPDATE savings_accounts
                SET balance_pesewas = balance_pesewas - ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (amount_pesewas, savings_account_id),
            )

        conn.commit()

        return get_payment_sheet_detail(conn, sheet_id)


@app.delete("/api/payment-sheets/{sheet_id}")
def delete_payment_sheet(sheet_id: int):
    with get_connection() as conn:
        sheet = conn.execute(
            "SELECT id, status FROM payment_sheets WHERE id = ?",
            (sheet_id,),
        ).fetchone()

        if not sheet:
            raise HTTPException(status_code=404, detail="Payment sheet not found.")

        entries = conn.execute(
            """
            SELECT id, bank_id, savings_account_id, amount_pesewas, source
            FROM payment_entries
            WHERE sheet_id = ?
            """,
            (sheet_id,),
        ).fetchall()

        if sheet["status"] == "posted":
            bank_groups = group_bank_payment_amounts(entries)
            savings_groups = group_savings_payment_amounts(entries)

            for bank_id, amount_pesewas in bank_groups.items():
                validate_bank_exists(conn, bank_id)
                conn.execute(
                    """
                    UPDATE banks
                    SET balance_pesewas = balance_pesewas + ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (amount_pesewas, bank_id),
                )

            for savings_account_id, amount_pesewas in savings_groups.items():
                validate_savings_account_exists(conn, savings_account_id)
                conn.execute(
                    """
                    UPDATE savings_accounts
                    SET balance_pesewas = balance_pesewas + ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    (amount_pesewas, savings_account_id),
                )

        conn.execute("DELETE FROM payment_entries WHERE sheet_id = ?", (sheet_id,))
        conn.execute("DELETE FROM payment_sheets WHERE id = ?", (sheet_id,))
        conn.commit()

    return {"ok": True, "message": "Payment sheet deleted successfully."}

def month_key(value: date) -> str:
    return value.strftime("%Y-%m")


def previous_month_key(value: date) -> str:
    year = value.year
    month = value.month - 1

    if month == 0:
        month = 12
        year -= 1

    return f"{year:04d}-{month:02d}"


def percentage_change(current: int, previous: int) -> float:
    if previous == 0 and current == 0:
        return 0

    if previous == 0:
        return 100

    return round(((current - previous) / previous) * 100, 2)


def movement_direction(current: int, previous: int) -> str:
    if current > previous:
        return "rising"

    if current < previous:
        return "falling"

    return "stable"


def serialize_report_money_row(row: sqlite3.Row):
    return {
        "name": row["name"],
        "total": pesewas_to_money(row["total"]),
    }


def serialize_ledger_row(row: sqlite3.Row):
    return {
        "category": row["category"],
        "subcategory": row["subcategory"],
        "entry_count": row["entry_count"],
        "total_amount": pesewas_to_money(row["total_amount"]),
        "first_date": row["first_date"],
        "last_date": row["last_date"],
    }


@app.get("/api/reports/summary")
def get_reports_summary(year: Optional[int] = None):
    report_year = year or date.today().year
    year_prefix = f"{report_year:04d}-%"

    today = date.today()
    current_month = month_key(today)
    previous_month = previous_month_key(today)

    with get_connection() as conn:
        total_receipts = conn.execute(
            """
            SELECT COALESCE(SUM(total_pesewas), 0) AS total
            FROM receipt_sheets
            WHERE status = 'posted'
            AND sheet_date LIKE ?
            """,
            (year_prefix,),
        ).fetchone()["total"]

        total_payments = conn.execute(
            """
            SELECT COALESCE(SUM(total_pesewas), 0) AS total
            FROM payment_sheets
            WHERE status = 'posted'
            AND sheet_date LIKE ?
            """,
            (year_prefix,),
        ).fetchone()["total"]

        bank_total = conn.execute(
            """
            SELECT COALESCE(SUM(balance_pesewas), 0) AS total
            FROM banks
            """
        ).fetchone()["total"]

        savings_total = conn.execute(
            """
            SELECT COALESCE(SUM(balance_pesewas), 0) AS total
            FROM savings_accounts
            """
        ).fetchone()["total"]

        petty_cash_balance = get_petty_cash_balance_pesewas(conn)

        income_breakdown_rows = conn.execute(
            """
            SELECT
                receipt_entries.category AS name,
                COALESCE(SUM(receipt_entries.amount_pesewas), 0) AS total
            FROM receipt_entries
            INNER JOIN receipt_sheets ON receipt_sheets.id = receipt_entries.sheet_id
            WHERE receipt_sheets.status = 'posted'
            AND receipt_sheets.sheet_date LIKE ?
            GROUP BY receipt_entries.category
            ORDER BY total DESC
            """,
            (year_prefix,),
        ).fetchall()

        expense_breakdown_rows = conn.execute(
            """
            SELECT
                payment_entries.category AS name,
                COALESCE(SUM(payment_entries.amount_pesewas), 0) AS total
            FROM payment_entries
            INNER JOIN payment_sheets ON payment_sheets.id = payment_entries.sheet_id
            WHERE payment_sheets.status = 'posted'
            AND payment_sheets.sheet_date LIKE ?
            GROUP BY payment_entries.category
            ORDER BY total DESC
            """,
            (year_prefix,),
        ).fetchall()

        receipt_month_rows = conn.execute(
            """
            SELECT
                substr(sheet_date, 1, 7) AS month,
                COALESCE(SUM(total_pesewas), 0) AS total
            FROM receipt_sheets
            WHERE status = 'posted'
            AND sheet_date LIKE ?
            GROUP BY substr(sheet_date, 1, 7)
            """,
            (year_prefix,),
        ).fetchall()

        payment_month_rows = conn.execute(
            """
            SELECT
                substr(sheet_date, 1, 7) AS month,
                COALESCE(SUM(total_pesewas), 0) AS total
            FROM payment_sheets
            WHERE status = 'posted'
            AND sheet_date LIKE ?
            GROUP BY substr(sheet_date, 1, 7)
            """,
            (year_prefix,),
        ).fetchall()

        receipt_month_map = {
            row["month"]: row["total"] for row in receipt_month_rows
        }

        payment_month_map = {
            row["month"]: row["total"] for row in payment_month_rows
        }

        monthly_trend = []

        for month_number in range(1, 13):
            key = f"{report_year:04d}-{month_number:02d}"
            receipts = receipt_month_map.get(key, 0)
            payments = payment_month_map.get(key, 0)

            monthly_trend.append(
                {
                    "month": key,
                    "receipts": pesewas_to_money(receipts),
                    "payments": pesewas_to_money(payments),
                    "net": pesewas_to_money(receipts - payments),
                }
            )

        def receipt_total_for_month(category: Optional[str], month: str):
            if category:
                row = conn.execute(
                    """
                    SELECT COALESCE(SUM(receipt_entries.amount_pesewas), 0) AS total
                    FROM receipt_entries
                    INNER JOIN receipt_sheets ON receipt_sheets.id = receipt_entries.sheet_id
                    WHERE receipt_sheets.status = 'posted'
                    AND receipt_sheets.sheet_date LIKE ?
                    AND receipt_entries.category = ?
                    """,
                    (f"{month}%", category),
                ).fetchone()
            else:
                row = conn.execute(
                    """
                    SELECT COALESCE(SUM(total_pesewas), 0) AS total
                    FROM receipt_sheets
                    WHERE status = 'posted'
                    AND sheet_date LIKE ?
                    """,
                    (f"{month}%",),
                ).fetchone()

            return row["total"]

        def payment_total_for_month(month: str):
            row = conn.execute(
                """
                SELECT COALESCE(SUM(total_pesewas), 0) AS total
                FROM payment_sheets
                WHERE status = 'posted'
                AND sheet_date LIKE ?
                """,
                (f"{month}%",),
            ).fetchone()

            return row["total"]

        current_tithe = receipt_total_for_month("Tithe", current_month)
        previous_tithe = receipt_total_for_month("Tithe", previous_month)

        current_receipts = receipt_total_for_month(None, current_month)
        previous_receipts = receipt_total_for_month(None, previous_month)

        current_payments = payment_total_for_month(current_month)
        previous_payments = payment_total_for_month(previous_month)

        comparisons = [
            {
                "name": "Tithe",
                "current_month": pesewas_to_money(current_tithe),
                "previous_month": pesewas_to_money(previous_tithe),
                "difference": pesewas_to_money(current_tithe - previous_tithe),
                "percentage_change": percentage_change(current_tithe, previous_tithe),
                "direction": movement_direction(current_tithe, previous_tithe),
            },
            {
                "name": "Total Receipts",
                "current_month": pesewas_to_money(current_receipts),
                "previous_month": pesewas_to_money(previous_receipts),
                "difference": pesewas_to_money(current_receipts - previous_receipts),
                "percentage_change": percentage_change(current_receipts, previous_receipts),
                "direction": movement_direction(current_receipts, previous_receipts),
            },
            {
                "name": "Total Payments",
                "current_month": pesewas_to_money(current_payments),
                "previous_month": pesewas_to_money(previous_payments),
                "difference": pesewas_to_money(current_payments - previous_payments),
                "percentage_change": percentage_change(current_payments, previous_payments),
                "direction": movement_direction(current_payments, previous_payments),
            },
        ]

        receipt_ledger_rows = conn.execute(
            """
            SELECT
                receipt_entries.category,
                receipt_entries.subcategory,
                COUNT(*) AS entry_count,
                COALESCE(SUM(receipt_entries.amount_pesewas), 0) AS total_amount,
                MIN(receipt_sheets.sheet_date) AS first_date,
                MAX(receipt_sheets.sheet_date) AS last_date
            FROM receipt_entries
            INNER JOIN receipt_sheets ON receipt_sheets.id = receipt_entries.sheet_id
            WHERE receipt_sheets.status = 'posted'
            AND receipt_sheets.sheet_date LIKE ?
            GROUP BY receipt_entries.category, receipt_entries.subcategory
            ORDER BY receipt_entries.category ASC, total_amount DESC
            """,
            (year_prefix,),
        ).fetchall()

        payment_ledger_rows = conn.execute(
            """
            SELECT
                payment_entries.category,
                payment_entries.subcategory,
                COUNT(*) AS entry_count,
                COALESCE(SUM(payment_entries.amount_pesewas), 0) AS total_amount,
                MIN(payment_sheets.sheet_date) AS first_date,
                MAX(payment_sheets.sheet_date) AS last_date
            FROM payment_entries
            INNER JOIN payment_sheets ON payment_sheets.id = payment_entries.sheet_id
            WHERE payment_sheets.status = 'posted'
            AND payment_sheets.sheet_date LIKE ?
            GROUP BY payment_entries.category, payment_entries.subcategory
            ORDER BY payment_entries.category ASC, total_amount DESC
            """,
            (year_prefix,),
        ).fetchall()

        bank_rows = conn.execute(
            """
            SELECT id, name, balance_pesewas, created_at, updated_at
            FROM banks
            ORDER BY name ASC
            """
        ).fetchall()

        savings_rows = conn.execute(
            """
            SELECT id, name, balance_pesewas, created_at, updated_at
            FROM savings_accounts
            ORDER BY name ASC
            """
        ).fetchall()

    total_assets = bank_total + savings_total + petty_cash_balance

    return {
        "year": report_year,
        "current_month": current_month,
        "previous_month": previous_month,
        "summary": {
            "total_receipts": pesewas_to_money(total_receipts),
            "total_payments": pesewas_to_money(total_payments),
            "net": pesewas_to_money(total_receipts - total_payments),
            "bank_total": pesewas_to_money(bank_total),
            "savings_total": pesewas_to_money(savings_total),
            "petty_cash_balance": pesewas_to_money(petty_cash_balance),
            "total_assets": pesewas_to_money(total_assets),
        },
        "comparisons": comparisons,
        "income_breakdown": [
            serialize_report_money_row(row) for row in income_breakdown_rows
        ],
        "expense_breakdown": [
            serialize_report_money_row(row) for row in expense_breakdown_rows
        ],
        "monthly_trend": monthly_trend,
        "receipt_ledger": [
            serialize_ledger_row(row) for row in receipt_ledger_rows
        ],
        "payment_ledger": [
            serialize_ledger_row(row) for row in payment_ledger_rows
        ],
        "balance_sheet": {
            "assets": {
                "banks": [serialize_bank(row) for row in bank_rows],
                "savings": [serialize_savings_account(row) for row in savings_rows],
                "petty_cash": pesewas_to_money(petty_cash_balance),
                "total_assets": pesewas_to_money(total_assets),
            }
        },
    }

@app.get("/api/reports/export-excel")
def export_report_excel(year: Optional[int] = None):
    report_year = year or date.today().year
    year_prefix = f"{report_year:04d}-%"

    report = get_reports_summary(report_year)

    reports_dir = BASE_DIR / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    file_path = reports_dir / f"church_finance_report_{report_year}.xlsx"

    workbook = Workbook()

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="0F172A")
    subtitle_font = Font(size=11, bold=True, color="475569")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    money_format = '"GHS" #,##0.00'

    def apply_header_style(sheet, row_number: int):
        for cell in sheet[row_number]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def apply_table_style(sheet, money_columns=None):
        money_columns = money_columns or []

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for column_index in money_columns:
            for row_number in range(2, sheet.max_row + 1):
                sheet.cell(row=row_number, column=column_index).number_format = money_format

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            sheet.column_dimensions[column_letter].width = min(max_length + 4, 45)

    def add_table(sheet, headers, rows, money_columns=None):
        sheet.append(headers)
        apply_header_style(sheet, 1)

        for row in rows:
            sheet.append(row)

        sheet.freeze_panes = "A2"
        apply_table_style(sheet, money_columns)

    # -------------------------
    # Summary Sheet
    # -------------------------
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    summary_sheet["A1"] = "Church Finance Report"
    summary_sheet["A1"].font = title_font

    summary_sheet["A2"] = f"Financial Year: {report_year}"
    summary_sheet["A2"].font = subtitle_font

    summary_rows = [
        ["Total Receipts", report["summary"]["total_receipts"]],
        ["Total Payments", report["summary"]["total_payments"]],
        ["Net Position", report["summary"]["net"]],
        ["Bank Total", report["summary"]["bank_total"]],
        ["Savings Total", report["summary"]["savings_total"]],
        ["Petty Cash Balance", report["summary"]["petty_cash_balance"]],
        ["Total Assets", report["summary"]["total_assets"]],
    ]

    summary_sheet.append([])
    summary_sheet.append(["Metric", "Amount"])

    for row in summary_rows:
        summary_sheet.append(row)

    apply_header_style(summary_sheet, 4)

    for row_number in range(5, 5 + len(summary_rows)):
        summary_sheet.cell(row=row_number, column=2).number_format = money_format

    apply_table_style(summary_sheet, [2])

    comparison_start = 14
    summary_sheet.cell(row=comparison_start, column=1, value="Month-to-Month Comparison")
    summary_sheet.cell(row=comparison_start, column=1).font = title_font

    comparison_headers = [
        "Name",
        "Previous Month",
        "Current Month",
        "Difference",
        "Change %",
        "Direction",
    ]

    for col_index, header in enumerate(comparison_headers, start=1):
        cell = summary_sheet.cell(row=comparison_start + 1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_index, item in enumerate(report["comparisons"], start=comparison_start + 2):
        summary_sheet.cell(row=row_index, column=1, value=item["name"])
        summary_sheet.cell(row=row_index, column=2, value=item["previous_month"])
        summary_sheet.cell(row=row_index, column=3, value=item["current_month"])
        summary_sheet.cell(row=row_index, column=4, value=item["difference"])
        summary_sheet.cell(row=row_index, column=5, value=item["percentage_change"])
        summary_sheet.cell(row=row_index, column=6, value=item["direction"])

        for col_index in [2, 3, 4]:
            summary_sheet.cell(row=row_index, column=col_index).number_format = money_format

    # -------------------------
    # Monthly Trend Sheet
    # -------------------------
    trend_sheet = workbook.create_sheet("Monthly Trend")
    trend_rows = [
        [
            row["month"],
            row["receipts"],
            row["payments"],
            row["net"],
        ]
        for row in report["monthly_trend"]
    ]

    add_table(
        trend_sheet,
        ["Month", "Receipts", "Payments", "Net"],
        trend_rows,
        money_columns=[2, 3, 4],
    )

    line_chart = LineChart()
    line_chart.title = "Monthly Receipts vs Payments"
    line_chart.y_axis.title = "Amount"
    line_chart.x_axis.title = "Month"

    data = Reference(trend_sheet, min_col=2, max_col=3, min_row=1, max_row=13)
    categories = Reference(trend_sheet, min_col=1, min_row=2, max_row=13)
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(categories)
    line_chart.height = 10
    line_chart.width = 20
    trend_sheet.add_chart(line_chart, "F2")

    bar_chart = BarChart()
    bar_chart.title = "Net Movement"
    bar_chart.y_axis.title = "Net Amount"
    bar_chart.x_axis.title = "Month"

    net_data = Reference(trend_sheet, min_col=4, min_row=1, max_row=13)
    bar_chart.add_data(net_data, titles_from_data=True)
    bar_chart.set_categories(categories)
    bar_chart.height = 10
    bar_chart.width = 20
    trend_sheet.add_chart(bar_chart, "F22")

    # -------------------------
    # Income Breakdown
    # -------------------------
    income_sheet = workbook.create_sheet("Income Breakdown")
    income_rows = [[row["name"], row["total"]] for row in report["income_breakdown"]]

    add_table(
        income_sheet,
        ["Category", "Total"],
        income_rows,
        money_columns=[2],
    )

    if len(income_rows) > 0:
        income_chart = PieChart()
        income_chart.title = "Income Breakdown"

        income_data = Reference(income_sheet, min_col=2, min_row=1, max_row=len(income_rows) + 1)
        income_labels = Reference(income_sheet, min_col=1, min_row=2, max_row=len(income_rows) + 1)

        income_chart.add_data(income_data, titles_from_data=True)
        income_chart.set_categories(income_labels)
        income_chart.height = 10
        income_chart.width = 15
        income_sheet.add_chart(income_chart, "D2")

    # -------------------------
    # Expense Breakdown
    # -------------------------
    expense_sheet = workbook.create_sheet("Expense Breakdown")
    expense_rows = [[row["name"], row["total"]] for row in report["expense_breakdown"]]

    add_table(
        expense_sheet,
        ["Category", "Total"],
        expense_rows,
        money_columns=[2],
    )

    if len(expense_rows) > 0:
        expense_chart = PieChart()
        expense_chart.title = "Expense Breakdown"

        expense_data = Reference(expense_sheet, min_col=2, min_row=1, max_row=len(expense_rows) + 1)
        expense_labels = Reference(expense_sheet, min_col=1, min_row=2, max_row=len(expense_rows) + 1)

        expense_chart.add_data(expense_data, titles_from_data=True)
        expense_chart.set_categories(expense_labels)
        expense_chart.height = 10
        expense_chart.width = 15
        expense_sheet.add_chart(expense_chart, "D2")

    # -------------------------
    # Receipt Ledger
    # -------------------------
    receipt_ledger_sheet = workbook.create_sheet("Receipt Ledger")
    receipt_ledger_rows = [
        [
            row["category"],
            row["subcategory"],
            row["entry_count"],
            row["total_amount"],
            row["first_date"],
            row["last_date"],
        ]
        for row in report["receipt_ledger"]
    ]

    add_table(
        receipt_ledger_sheet,
        ["Category", "Field", "Entries", "Total", "First Date", "Last Date"],
        receipt_ledger_rows,
        money_columns=[4],
    )

    # -------------------------
    # Payment Ledger
    # -------------------------
    payment_ledger_sheet = workbook.create_sheet("Payment Ledger")
    payment_ledger_rows = [
        [
            row["category"],
            row["subcategory"],
            row["entry_count"],
            row["total_amount"],
            row["first_date"],
            row["last_date"],
        ]
        for row in report["payment_ledger"]
    ]

    add_table(
        payment_ledger_sheet,
        ["Category", "Field", "Entries", "Total", "First Date", "Last Date"],
        payment_ledger_rows,
        money_columns=[4],
    )

    with get_connection() as conn:
        receipt_details = conn.execute(
            """
            SELECT
                receipt_sheets.sheet_date,
                receipt_sheets.title AS sheet_title,
                receipt_entries.category,
                receipt_entries.subcategory,
                receipt_entries.amount_pesewas,
                banks.name AS bank_name,
                receipt_entries.note
            FROM receipt_entries
            INNER JOIN receipt_sheets ON receipt_sheets.id = receipt_entries.sheet_id
            INNER JOIN banks ON banks.id = receipt_entries.bank_id
            WHERE receipt_sheets.status = 'posted'
            AND receipt_sheets.sheet_date LIKE ?
            ORDER BY receipt_sheets.sheet_date ASC, receipt_entries.id ASC
            """,
            (year_prefix,),
        ).fetchall()

        payment_details = conn.execute(
            """
            SELECT
                payment_sheets.sheet_date,
                payment_sheets.title AS sheet_title,
                payment_entries.category,
                payment_entries.subcategory,
                payment_entries.amount_pesewas,
                payment_entries.source,
                CASE
                    WHEN payment_entries.source = 'bank' THEN banks.name
                    WHEN payment_entries.source = 'savings' THEN savings_accounts.name
                    ELSE 'Petty Cash'
                END AS account_name,
                payment_entries.note
            FROM payment_entries
            INNER JOIN payment_sheets ON payment_sheets.id = payment_entries.sheet_id
            LEFT JOIN banks ON banks.id = payment_entries.bank_id
            LEFT JOIN savings_accounts ON savings_accounts.id = payment_entries.savings_account_id
            WHERE payment_sheets.status = 'posted'
            AND payment_sheets.sheet_date LIKE ?
            ORDER BY payment_sheets.sheet_date ASC, payment_entries.id ASC
            """,
            (year_prefix,),
        ).fetchall()

        savings_transfers = conn.execute(
            """
            SELECT
                savings_transfers.created_at,
                banks.name AS bank_name,
                savings_accounts.name AS savings_account_name,
                savings_transfers.amount_pesewas,
                savings_transfers.note
            FROM savings_transfers
            INNER JOIN banks ON banks.id = savings_transfers.bank_id
            INNER JOIN savings_accounts ON savings_accounts.id = savings_transfers.savings_account_id
            WHERE savings_transfers.created_at LIKE ?
            ORDER BY savings_transfers.created_at ASC
            """,
            (year_prefix,),
        ).fetchall()

        petty_cash_withdrawals = conn.execute(
            """
            SELECT
                petty_cash_withdrawals.created_at,
                banks.name AS bank_name,
                petty_cash_withdrawals.amount_pesewas,
                petty_cash_withdrawals.note
            FROM petty_cash_withdrawals
            INNER JOIN banks ON banks.id = petty_cash_withdrawals.bank_id
            WHERE petty_cash_withdrawals.created_at LIKE ?
            ORDER BY petty_cash_withdrawals.created_at ASC
            """,
            (year_prefix,),
        ).fetchall()

    # -------------------------
    # Receipt Details
    # -------------------------
    receipt_details_sheet = workbook.create_sheet("Receipt Details")
    receipt_detail_rows = [
        [
            row["sheet_date"],
            row["sheet_title"],
            row["category"],
            row["subcategory"],
            pesewas_to_money(row["amount_pesewas"]),
            row["bank_name"],
            row["note"] or "",
        ]
        for row in receipt_details
    ]

    add_table(
        receipt_details_sheet,
        ["Date", "Sheet", "Category", "Field", "Amount", "Bank", "Note"],
        receipt_detail_rows,
        money_columns=[5],
    )

    # -------------------------
    # Payment Details
    # -------------------------
    payment_details_sheet = workbook.create_sheet("Payment Details")
    payment_detail_rows = [
        [
            row["sheet_date"],
            row["sheet_title"],
            row["category"],
            row["subcategory"],
            pesewas_to_money(row["amount_pesewas"]),
            row["source"],
            row["account_name"],
            row["note"] or "",
        ]
        for row in payment_details
    ]

    add_table(
        payment_details_sheet,
        ["Date", "Sheet", "Category", "Field", "Amount", "Source", "Account", "Note"],
        payment_detail_rows,
        money_columns=[5],
    )

    # -------------------------
    # Balance Sheet
    # -------------------------
    balance_sheet = workbook.create_sheet("Balance Sheet")

    balance_rows = []

    for bank in report["balance_sheet"]["assets"]["banks"]:
        balance_rows.append(["Bank", bank["name"], bank["balance"]])

    for savings in report["balance_sheet"]["assets"]["savings"]:
        balance_rows.append(["Savings", savings["name"], savings["balance"]])

    balance_rows.append(["Petty Cash", "Petty Cash", report["summary"]["petty_cash_balance"]])
    balance_rows.append(["Total Assets", "Total Assets", report["summary"]["total_assets"]])

    add_table(
        balance_sheet,
        ["Asset Type", "Account", "Amount"],
        balance_rows,
        money_columns=[3],
    )

    # -------------------------
    # Savings Transfers
    # -------------------------
    savings_transfer_sheet = workbook.create_sheet("Savings Transfers")
    savings_transfer_rows = [
        [
            row["created_at"],
            row["bank_name"],
            row["savings_account_name"],
            pesewas_to_money(row["amount_pesewas"]),
            row["note"] or "",
        ]
        for row in savings_transfers
    ]

    add_table(
        savings_transfer_sheet,
        ["Date", "From Bank", "To Savings", "Amount", "Note"],
        savings_transfer_rows,
        money_columns=[4],
    )

    # -------------------------
    # Petty Cash Withdrawals
    # -------------------------
    petty_sheet = workbook.create_sheet("Petty Cash Withdrawals")
    petty_rows = [
        [
            row["created_at"],
            row["bank_name"],
            pesewas_to_money(row["amount_pesewas"]),
            row["note"] or "",
        ]
        for row in petty_cash_withdrawals
    ]

    add_table(
        petty_sheet,
        ["Date", "Bank", "Amount", "Note"],
        petty_rows,
        money_columns=[3],
    )

    workbook.save(file_path)

    return FileResponse(
        path=file_path,
        filename=file_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/api/dashboard/summary")
def dashboard_summary():
    with get_connection() as conn:
        bank_summary = conn.execute(
            """
            SELECT
                COUNT(*) AS bank_count,
                COALESCE(SUM(balance_pesewas), 0) AS total_balance
            FROM banks
            """
        ).fetchone()

        savings_total = conn.execute(
            """
            SELECT COALESCE(SUM(balance_pesewas), 0) AS total
            FROM savings_accounts
            """
        ).fetchone()["total"]

        receipts_total = conn.execute(
            """
            SELECT COALESCE(SUM(total_pesewas), 0) AS total
            FROM receipt_sheets
            WHERE status = 'posted'
            """
        ).fetchone()["total"]

        payments_total = conn.execute(
            """
            SELECT COALESCE(SUM(total_pesewas), 0) AS total
            FROM payment_sheets
            WHERE status = 'posted'
            """
        ).fetchone()["total"]

        petty_cash_balance = get_petty_cash_balance_pesewas(conn)

    return {
        "bank_count": bank_summary["bank_count"],
        "total_balance": pesewas_to_money(bank_summary["total_balance"]),
        "receipts_total": pesewas_to_money(receipts_total),
        "payments_total": pesewas_to_money(payments_total),
        "petty_cash_balance": pesewas_to_money(petty_cash_balance),
        "savings_total": pesewas_to_money(savings_total),
    }